import requests
import urllib
import argparse
import yaml
import openpyxl
import logging
import os

class MyShop(object):
    IMAGE_URL = "/admin/index.php?route=api/images/{name}"
    PRODUCT_URL = "/admin/index.php?route=api/products/{name}"
    CATEGORY_URL = "/admin/index.php?route=api/categories/{name}"
    MANUFACTURER_URL = "/admin/index.php?route=api/manufacturers/{name}"
    STATE_FILE = 'import.state'
    HEADERS = {
        'name': 'Name',
        'model': 'Model',
        'description': 'Description',
        'images': 'Image Paths',
        'meta_title': 'Meta Title',
        'meta_description': 'Meta Description',
        'meta_keyword': 'Meta Keywords',
        'manufacturer': 'Manufacturer',
        'requires_shipping': 'Requires Shipping',
        'weight': "Weight",
        'price': "Price",
        'weight_unit': "Weight Unit",
        'length': "Length",
        'width': "Width",
        'height': 'Height',
        'dimension_unit': "Dimension Unit",
        'status': "Status",
        'sku': "SKU",
        'upc': "UPC",
        'ean': "EAN",
        'jan': "JAN",
        'isbn': "ISBN",
        'mpn': "MPN",
        'location': "Location",
        'points': "Points",
        'sort_order': 'Sort Order',
        'tax_class_id': "Tax Class ID",
        'stock_status_id': "Stock Status ID",
        'categories': "Category Paths",
        'seo_name': "SEO Name"
    }

    LIST_TYPES = ['images', 'categories']

    def __init__(self, options):
        self.options = options
        with open(options.config, 'r') as c:
            self._config = yaml.load(c, Loader=yaml.Loader)
        self.login(self._config['store_url'], self._config['api_user'], self._config['api_key'])

    def _handle_error(self, ret):
        if ret.status_code == 200:
            try:
                return ret.json()
            except:
                return ret.text
        else:
            raise Exception(dict(code=ret.status_code, msg=ret.text))

    def _get_url(self, uri, name):
        name = urllib.parse.quote_plus(name)
        return self._handle_error(requests.get(self._store_url + uri.format(name=name),
                                        headers=dict(
                                            Authorization="Bearer " + self._token
                                            ),
                                        ))

    def _delete_url(self, uri, name):
        name = urllib.parse.quote_plus(name)
        return self._handle_error(requests.delete(self._store_url + uri.format(name=name),
                                        headers=dict(
                                            Authorization="Bearer " + self._token
                                            ),
                                        ))

    def _post_url(self, uri, data):
        return self._handle_error(requests.post(self._store_url + uri.replace('/{name}', ''),
                                json=data,
                                headers=dict(
                                    Authorization="Bearer " + self._token
                                    ),
                                ))

    def load_from_db(self, db):
        record = db(db.myshop_configurations.created_by == self._user_id).select().first()
        if record:
            self._store_url = record.store_url
            self.login(record.store_url, record.api_user, record.api_key)
        
    @property
    def token(self):
        return self._token

    def login(self, store_url, api_user, api_key):
        self._store_url = store_url
        ret = requests.post(store_url + '/index.php?route=api/login',json={'username':api_user, 'key':api_key})
        jsn = ret.json()
        if ret.status_code == 200 and 'api_token' in jsn:
            self._token = jsn['api_token']
        else:
            print(jsn)
            self._token = None

    def get_product(self, name):
        return self._get_url(self.PRODUCT_URL, name)

    def get_products(self):
        return self.get_product('all')

    def delete_product(self, name):
        return self._delete_url(self.PRODUCT_URL, name)

    def post_product(self, data):
        return self._post_url(self.PRODUCT_URL, data)

    def get_category(self, name):
        return self._get_url(self.CATEGORY_URL, name)

    def delete_category(self, name):
        return self._delete_url(self.CATEGORY_URL, name)

    def post_category(self, data):
        return self._post_url(self.CATEGORY_URL, data)

    def get_manufacturer(self, name):
        return self._get_url(self.MANUFACTURER_URL, name)

    def delete_manufacturer(self, name):
        return self._delete_url(self.MANUFACTURER_URL, name)

    def post_manufacturer(self, data):
        return self._post_url(self.MANUFACTURER_URL, data)

    def content_type(self, filename):
        import mimetypes
        mimetypes.init()
        m_type =  mimetypes.guess_type(filename)[0]
        if m_type not in ('image/jpeg', 'image/png'):
            raise Exception(filename + " is not support image type")
        return m_type

    def _post_image(self, src_stream, content_type, dst):
        return self._handle_error(requests.post(self._store_url + self.IMAGE_URL.format(name=dst),
                    data=src_stream,
                    headers={
                        "Authorization": "Bearer " + self._token,
                        "content-type": content_type
                        },
                    ))

    def post_image(self, src, dst, src_stream=None):
        if src_stream == None:
            with open(src, 'rb') as src_stream:
                self._post_image(src_stream, self.content_type(src), dst)
        else:
            self._post_image(src_stream, self.content_type(src), dst)

    def _get_product_row(self, p_detail):
        row = [''] * len(self._headers)
        for (k, v) in p_detail.items():
            if k in self._attr:
                if k in self.LIST_TYPES:
                    v = ','.join(v)
                row[self._attr[k]] = v
        return row

    def _get_xls_headers(self):
        self._attr = dict()
        row = []
        i = 0
        for (k, v) in self.HEADERS.items():
            row.append(v)
            self._attr[k] = i
            i = i + 1
        self._headers = row
        return row
        
    def export_products(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(self._get_xls_headers())
        for product in s.get_products():
            try:
                sheet.append(self._get_product_row(s.get_product(product)))
            except Exception as e:
                logging.error("{}: {}".format(product, str(e)))
        workbook.save(self.options.products)
        workbook.close()

    def _parse_header(self, row):
        headers = { value: key for key, value in self.HEADERS.items() }
        self._attr = dict()
        i = 0
        for c in row:
            if c.value in headers:
                self._attr[i] = headers[c.value]
                i += 1

    def _get_product_from_row(self, row):
        product = dict()
        i = 0
        for c in row:
            if self._attr[i] in self.LIST_TYPES:
                if c.value != None and len(c.value) > 0:
                    val = c.value.split(',')
                else:
                    val = []
            else:
                val = c.value
            product[self._attr[i]] = val
            i += 1
        return product
            

    def import_products(self):
        workbook = openpyxl.load_workbook(self.options.products)
        sheet = workbook.active
        header = True
        if os.path.isfile(self.STATE_FILE) and self.options.maintain_state:
            start_import = False
            with open(self.STATE_FILE, 'r') as st_file:
                last_imported_product = st_file.read()
        else:
            start_import = True

        for row in sheet.iter_rows():
            if header:
                self._parse_header(row)
                header = False
            else:
                obj = self._get_product_from_row(row)

                if not start_import:
                    if last_imported_product == obj['name']:
                        start_import = True
                    continue

                if 'images' in obj and not options.skip_images:
                    for img in obj['images']:
                        try:
                            self.post_image(img, img)
                        except Exception as e:
                            logging.error("Updating image '{}' for product name '{}', Message: {}".format(img, obj['name'], str(e)))
                            return
                try:
                    self.post_product(obj)
                    if self.options.maintain_state:
                        with open(self.STATE_FILE, 'w') as st_file:
                            st_file.write(obj['name'])
                except Exception as e:
                    raise e

                
            

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Utility to bulk create/update products, categories & manufacturers in your myshop store.')
    parser.add_argument('-c', '--config', default='config.yaml', help='Store configuration file')
    parser.add_argument('-p', '--products', default='products.xlsx', help='File that will contain the products informations')
    parser.add_argument('-s', '--skip-images', default=False, action='store_true', help='Skip importing image files associated with the products')
    parser.add_argument('-e', '--products_export', default=False, action='store_true', help='Export products from myshop store to local files')
    parser.add_argument('-i', '--products_import', default=False, action='store_true', help='Import products to myshop store')
    parser.add_argument('-m', '--maintain_state', default=False, action='store_true', help='Maintain import state so that the next import runs from where it stopped')
    logging.basicConfig(format='%(asctime)s %(levelname)-8s %(message)s',
                         level=logging.DEBUG,
                         datefmt='%Y-%m-%d %H:%M:%S')
    options = parser.parse_args()
    try:
        s = MyShop(options)
        if options.products_export:
            s.export_products()
        elif options.products_import:
            s.import_products()
        else:
            parser.print_usage()
    except Exception as e:
        logging.error(str(e))
